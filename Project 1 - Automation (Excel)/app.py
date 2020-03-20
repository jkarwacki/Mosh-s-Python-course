import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):git
    # function that takes in a spreadsheet and manipulates it, in this case multiplying all values on column 3 w/ 0.9
    # and adding the resulting values in column 4, then saving the file
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
